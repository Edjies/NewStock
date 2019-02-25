# -*-coding:utf-8 -*-
import requests
import json
import os
import numpy as np
import StockConfig
import StockIO
from tkinter import messagebox
import time
import datetime
import StockIndicator
from numpy import genfromtxt
from openpyxl import load_workbook

track_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir, os.pardir, os.pardir, 'stock', 'track'))
# 便宜的股票要拿得住，一定要留有底仓

def get_track_data():
    # data = []
    # with open('{}/{}'.format(StockConfig.path_track, '2_sma_track.txt'), 'r', encoding='utf-8') as f:
    #     for line in f.readlines():
    #         if not line.startswith("#") and not '\n' == line:
    #             data.append(line.strip('\n').split(','))
    #
    # # with open('{}/{}'.format(StockConfig.path_track, '1_sma_track.txt'), 'r', encoding='utf-8') as f:
    # #     for line in f.readlines():
    # #         if not line.startswith("#") and not '\n' == line:
    # #             data.append(line.strip('\n').split(','))
    #
    # data = np.array(data)

    return read_xlsx('data/stock.xlsx')


def read_xlsx(filename):
    wb = load_workbook(filename=filename)
    array = []
    sheet = wb['Sheet1']
    print(sheet['A1'].value)
    for row in sheet.iter_rows(min_row=2, max_col=8, max_row=1000):
        if row[0].value is None:
            break
        item = []
        for index in range(0, 8):
            item.append('' if row[index].value is None else row[index].value)
        array.append(item)
    npdata = np.array(array, dtype='str')
    print(npdata)
    return npdata



def track():
    """
    :param targets: 需要追踪的目标集合
    @:param code:
    @:param price:
    :return:
    """
    # 读取本地数据
    track_data = get_track_data()
    kline_map = StockIO.get_kline_map(track_data[:, 0], StockConfig.kline_type_day)

    page = 0
    pageNum = 40
    result = []
    code_list = track_data[:, 0]
    while page * pageNum < len(code_list):
        from_position = page * pageNum
        to_position = min((page + 1) * pageNum, len(code_list))
        data = code_list[from_position:to_position]
        cur_track_data = track_data[from_position:to_position]
        page += 1
        # 请求网络数据
        codes = ''
        for code in data:
            if code.startswith('6'):
                code = '0' + code
            else:
                code = '1' + code
            codes = codes + code + ','

        session = requests.Session()
        session.trust_env = False
        url = 'http://api.money.126.net/data/feed/{}'.format(codes)[:-1]
        print(url)
        r = session.get(url)
        quote = json.loads(r.text[len('_ntes_quote_callback('): -2])

        # 分析数据
        for target in cur_track_data:
            target_code = target[0]
            target_sma_down = 0 if target[2].strip() == '' else int(target[2])
            target_sma_up = 0 if target[3].strip() == '' else int(target[3])
            target_price_down = 0 if target[4].strip() == '' else float(target[4])
            target_price_up = 0 if target[5].strip() == '' else float(target[5])
            target_chg_down = 0 if target[6].strip() == '' else float(target[6])
            target_chg_up = 0 if target[7].strip() == '' else float(target[7])
            if target_code.startswith('6'):
                target_code = '0' + target_code
            else:
                target_code = '1' + target_code
            new_date = datetime.datetime.strptime(quote[target_code]['time'], '%Y/%m/%d %H:%M:%S').strftime('%Y-%m-%d')
            kline = kline_map[target_code[1:]]
            old_kline_item = kline[-1]
            new_kline_item = np.array([new_date, quote[target_code]['open'], old_kline_item[-3], quote[target_code]['high'], quote[target_code]['low'], quote[target_code]['volume']])
            if(new_kline_item[-1] == old_kline_item[-1]):
                kline[-1] = new_kline_item
            else:
                kline = np.row_stack((kline, new_kline_item))

            #print(kline[-1])

            # 跌破目标均线
            if target_sma_down != 0:
                message = ''
                sma, = StockIndicator.sma(kline, abs(target_sma_down))
                price = quote[target_code]['price']
                if (price < sma[-1]):
                    print(sma[-1], price)
                    print(r.text)
                    message = target_code[1:] + '跌破{}日线:'.format(abs(target_sma_down))
                if message != '':
                    messagebox.showinfo("tips", message)

            # 突破目标均线
            if target_sma_up != 0:
                message = ''
                sma, = StockIndicator.sma(kline, abs(target_sma_up))
                price = quote[target_code]['low']
                if (price > sma[-1]):
                    message = target_code[1:] + '突破{}日线:'.format(abs(target_sma_up))

                if message != '':
                    messagebox.showinfo("tips", message)

            # 跌到目标价位
            if target_price_down != 0:
                message = ''
                cur_price = float(quote[target_code]['low'])
                message = ''
                if cur_price <= abs(target_price_down):
                    message = target_code[1:] + '跌到目标价位:' + str(abs(target_price_down))
                if message != '':
                    messagebox.showinfo("tips", message)

            # 涨到目标价位
            if target_price_up != 0:
                message = ''
                cur_price = float(quote[target_code]['price'])
                message = ''
                if cur_price >= abs(target_price_up):
                    message = target_code[1:] + '涨到目标价位:' + str(abs(target_price_up))
                if message != '':
                    messagebox.showinfo("tips", message)

            # 每日目标跌幅
            if target_chg_down != 0:
                p_close = float(quote[target_code]['yestclose'])
                low = float(quote[target_code]['low'])
                chg = np.round(((low - p_close) / p_close * 100), decimals=2)
                if chg < target_chg_down:
                    message = target_code[1:] + '跌到目标跌幅:' + str(chg)
                    if message != '':
                        messagebox.showinfo("tips", message)

            if target_chg_up != 0:
                p_close = float(quote[target_code]['yestclose'])
                high = float(quote[target_code]['high'])
                chg = np.round(((high - p_close) / p_close * 100), decimals=2)
                if chg > target_chg_up:
                    message = target_code[1:] + '涨到目标涨幅:' + str(chg)
                    if message != '':
                        messagebox.showinfo("tips", message)
            # 每日目标涨幅


            # 默认会跟踪自动筛选出来的 sma变化
            # if target_sma_down == 0 and target_price_down == 0 and target_sma_up == 0 and target_price_up == 0:
            #     sma5, sma10, sma20, sma30 = StockIndicator.sma(kline, 5, 10, 20, 30)
            #     price = quote[target_code]['price']
            #     # # sma5 条件
            #     # if price < sma5[-1]:
            #     #    message = target_code[1:] + '跌破{}日线:'.format(5)
            #     # sma10 条件
            #     if price < sma20[-1]:
            #         message = target_code[1:] + '跌破{}日线:'.format(20)
            #     # sma20 条件
            #     # elif price < sma20[-1]:
            #     #     message = target_code[1:] + '跌破{}日线:'.format(20)
            #     else:
            #         message = ''
            #     if message != '':
            #         messagebox.showinfo("tips", message)
    return result

    # tk显示结果


def startTrack():
    while True:
        try:
            print(track())

        except Exception as e:
            print(e)
            pass
        time.sleep(10)


if __name__=="__main__":
    startTrack()
    # 如果当前时间下午15点，则中止

#