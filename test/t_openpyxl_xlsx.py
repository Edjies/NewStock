# -*- coding:utf-8 -*-
from numpy import genfromtxt
from openpyxl import load_workbook
import numpy


def read_xlsx():
    wb = load_workbook(filename='stock.xlsx')
    array = []
    sheet = wb['Sheet1']
    print(sheet['A1'].value)
    for row in sheet.iter_rows(min_row=2, max_col=8, max_row=1000):
        if row[0].value is None:
            break
        item = []
        for index in range(0, 8):
            item.append('' if row[index].value is None else row[index].value)
        array.append(item)
    npdata = numpy.array(array, dtype='str')
    print(npdata)
    pass


if __name__=='__main__':
    read_xlsx()
    pass

